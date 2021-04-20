from datetime import datetime
from openpyxl import load_workbook
import mapeo
from tarea import Tarea
import logging

logger = logging.getLogger('logGeneral')

# archivo GEOTASK
def importarArchivoGEotask(archivoGE): 

    logger.debug("Entra a funcionesXLSX.importarArchivoGEotask")

    workbookGE  = load_workbook(filename=archivoGE, read_only=True)
    sheetGE     = workbookGE.active
    claseGE     = []
    errorMsg    = ''

    # Using the values_only because you just want to return the cell value
    for row in sheetGE.iter_rows(values_only=True):

        logger.debug(row)

        if row[mapeo.FECHA_GE] == 'Fecha':
            pass

        else :
            # You need to parse the date from the spreadsheet into a datetime format
            fecha = row[mapeo.FECHA_GE]
            logger.debug("Fecha del archivo: %s", fecha)
            
            fechaHoraGE = datetime.strptime(fecha, "%Y/%m/%d")
            fechaGE = fechaHoraGE.date()

            tareaGE = Tarea(row[mapeo.PROYECTO_GE],
                            row[mapeo.TAREA_GE],
                            fechaGE,
                            row[mapeo.TIEMPO_GE])
            claseGE.append(tareaGE)

    primerObjeto =  claseGE[0].imprimir()
    logger.debug("Primer objeto: " + primerObjeto)

    logger.debug("Sale de funcionesXLSX.importarArchivoGEotask")
    return claseGE, errorMsg

#archivo CLICKUP
def importarArchivoCLickup(archivoCL, tarea_combo): 

    logger.debug("Entra a funcionesXLSX.importarArchivoCLickup")

    workbookCL  = load_workbook(filename=archivoCL, read_only=True)
    sheetCL     = workbookCL.active
    claseCL     = []
    errorMsg    = ''

    for row in sheetCL.iter_rows(values_only=True):

        logger.debug(row)

        if row[mapeo.FECHA_CL] == 'Start Text':
            pass

        else :
            # You need to parse the date from the spreadsheet into a datetime format
            fecha = row[mapeo.FECHA_CL]
            logger.debug("Fecha del archivo: %s", fecha)
            
            fechaHoraCL = datetime.strptime(fecha, "%m/%d/%Y, %H:%M:%S %p %Z")
            fechaCL = fechaHoraCL.date()

            tiempoCL = (row[mapeo.TIEMPO_CL] / 3600000)

            tareaCL = Tarea(row[mapeo.PROYECTO_CL],
                        row[mapeo.TAREA_CL],
                        fechaCL,
                        tiempoCL)
            claseCL.append(tareaCL)

    primerObjeto =  claseCL[0].imprimir()
    logger.debug("Primer objeto: " + primerObjeto)

    logger.debug("Sale de funcionesXLSX.importarArchivoCLickup")
    return claseCL, errorMsg