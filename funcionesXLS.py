import pandas as pd

archivo_excel = pd.read_excel('/ruta/ExcelEjemplo.xlsx')

print(archivo_excel.columns)
values = archivo_excel['Identificador'].values
print(values)
columnas = ['Identificador', 'Nombre', 'Apellidos']
df_seleccionados = archivo_excel[columnas]
print(df_seleccionados)



# --------------------------------------------
from datetime import datetime
from openpyxl import load_workbook
import mapeo
import tarea

# archivo GEOTASK
def importarArchivoGEotask(archivoGE, claseGE, errorMsg): 

    workbookGE  = load_workbook(filename=archivoGE, read_only=True)
    sheetGE     = workbookGE.active
    claseGE     = []
    errorMsg    = ''

    # Using the values_only because you just want to return the cell value
    for row in sheetGE.iter_rows(values_only=True):

        if row[mapeo.FECHA_GE] == 'Fecha':
            pass

        else :
            # You need to parse the date from the spreadsheet into a datetime format
            fecha = row[mapeo.FECHA_GE]
            print(fecha)
            fechaGE = datetime.strptime(fecha, "%Y/%m/%d")

            tareaGE = tarea.Tarea(proyecto=row[mapeo.PROYECTO_GE],
                            tarea=row[mapeo.TAREA_GE],
                            fecha=fechaGE,
                            tiempo=row[mapeo.TIEMPO_GE])
            claseGE.append(tareaGE)
    print(claseGE[0])


#archivo CLICKUP
def importarArchivoCLickup(archivoCL, claseCL, errorMsg): 

    workbookCL  = load_workbook(filename=archivoCL, read_only=True)
    sheetCL     = workbookCL.active
    claseCL     = []
    errorMsg    = ''

    for row in sheetCL.iter_rows(values_only=True):

        # You need to parse the date from the spreadsheet into a datetime format
        fecha = row[mapeo.FECHA_CL]
        fechaCL = datetime.strptime(fecha, "%Y-%m-%d")

        tareaCL = tarea.Tarea(proyecto=row[mapeo.PROYECTO_CL],
                        tarea=row[mapeo.TAREA_CL],
                        fecha=fechaCL,
                        tiempo=row[mapeo.TIEMPO_CL])
        claseCL.append(tareaCL)

    print(claseCL[0])

